from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from database import get_db, init_db
from config import TALLER
import io
import os

app = Flask(__name__)
app.secret_key = 'taller_secreto_2024'

# ─── INIT ───────────────────────────────────────────────────────────────────

@app.before_request
def setup():
    if not hasattr(app, '_db_initialized'):
        init_db()
        app._db_initialized = True

# ─── INDEX ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    db = get_db()
    stats = {
        'clientes': db.execute('SELECT COUNT(*) FROM clientes').fetchone()[0],
        'vehiculos': db.execute('SELECT COUNT(*) FROM vehiculos').fetchone()[0],
        'ordenes_abiertas': db.execute("SELECT COUNT(*) FROM ordenes WHERE estado='abierta'").fetchone()[0],
        'ordenes_cerradas': db.execute("SELECT COUNT(*) FROM ordenes WHERE estado='cerrada'").fetchone()[0],
    }
    ordenes_recientes = db.execute('''
        SELECT o.id, o.fecha_ingreso, o.estado,
               c.nombre || ' ' || c.apellido AS cliente,
               v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS vehiculo
        FROM ordenes o
        JOIN clientes c ON o.cliente_id = c.id
        JOIN vehiculos v ON o.vehiculo_id = v.id
        ORDER BY o.created_at DESC LIMIT 5
    ''').fetchall()
    db.close()
    return render_template('index.html', stats=stats, ordenes_recientes=ordenes_recientes)

# ─── CLIENTES ───────────────────────────────────────────────────────────────

@app.route('/clientes')
def clientes():
    q = request.args.get('q', '')
    db = get_db()
    if q:
        rows = db.execute(
            "SELECT * FROM clientes WHERE nombre LIKE ? OR apellido LIKE ? OR telefono LIKE ? ORDER BY apellido",
            (f'%{q}%', f'%{q}%', f'%{q}%')
        ).fetchall()
    else:
        rows = db.execute('SELECT * FROM clientes ORDER BY apellido').fetchall()
    db.close()
    return render_template('clientes.html', clientes=rows, q=q)

@app.route('/clientes/nuevo', methods=['GET', 'POST'])
def nuevo_cliente():
    if request.method == 'POST':
        nombre = request.form['nombre'].strip()
        apellido = request.form['apellido'].strip()
        direccion = request.form.get('direccion', '').strip()
        telefono = request.form.get('telefono', '').strip()
        email = request.form.get('email', '').strip()
        dni = request.form.get('dni', '').strip()
        if not nombre or not apellido:
            flash('Nombre y apellido son obligatorios.', 'danger')
        else:
            db = get_db()
            db.execute('INSERT INTO clientes (nombre,apellido,direccion,telefono,email,dni) VALUES (?,?,?,?,?,?)',
                       (nombre, apellido, direccion, telefono, email, dni))
            db.commit()
            db.close()
            flash('Cliente agregado correctamente.', 'success')
            return redirect(url_for('clientes'))
    return render_template('cliente_form.html', cliente=None)

@app.route('/clientes/<int:id>/editar', methods=['GET', 'POST'])
def editar_cliente(id):
    db = get_db()
    cliente = db.execute('SELECT * FROM clientes WHERE id=?', (id,)).fetchone()
    if not cliente:
        db.close()
        flash('Cliente no encontrado.', 'danger')
        return redirect(url_for('clientes'))
    if request.method == 'POST':
        nombre = request.form['nombre'].strip()
        apellido = request.form['apellido'].strip()
        direccion = request.form.get('direccion', '').strip()
        telefono = request.form.get('telefono', '').strip()
        email = request.form.get('email', '').strip()
        dni = request.form.get('dni', '').strip()
        db.execute('UPDATE clientes SET nombre=?,apellido=?,direccion=?,telefono=?,email=?,dni=? WHERE id=?',
                   (nombre, apellido, direccion, telefono, email, dni, id))
        db.commit()
        db.close()
        flash('Cliente actualizado.', 'success')
        return redirect(url_for('clientes'))
    db.close()
    return render_template('cliente_form.html', cliente=cliente)

@app.route('/clientes/<int:id>')
def ver_cliente(id):
    db = get_db()
    cliente = db.execute('SELECT * FROM clientes WHERE id=?', (id,)).fetchone()
    vehiculos = db.execute('SELECT * FROM vehiculos WHERE cliente_id=?', (id,)).fetchall()
    ordenes = db.execute('''
        SELECT o.*, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS vehiculo
        FROM ordenes o JOIN vehiculos v ON o.vehiculo_id=v.id
        WHERE o.cliente_id=? ORDER BY o.fecha_ingreso DESC
    ''', (id,)).fetchall()
    db.close()
    return render_template('cliente_detalle.html', cliente=cliente, vehiculos=vehiculos, ordenes=ordenes)

@app.route('/clientes/<int:id>/eliminar', methods=['POST'])
def eliminar_cliente(id):
    db = get_db()
    db.execute('DELETE FROM clientes WHERE id=?', (id,))
    db.commit()
    db.close()
    flash('Cliente eliminado.', 'success')
    return redirect(url_for('clientes'))

# ─── VEHÍCULOS ──────────────────────────────────────────────────────────────

@app.route('/vehiculos')
def vehiculos():
    q = request.args.get('q', '')
    db = get_db()
    if q:
        rows = db.execute('''
            SELECT v.*, c.nombre || ' ' || c.apellido AS cliente
            FROM vehiculos v LEFT JOIN clientes c ON v.cliente_id=c.id
            WHERE v.marca LIKE ? OR v.modelo LIKE ? OR v.patente LIKE ?
            ORDER BY v.marca
        ''', (f'%{q}%', f'%{q}%', f'%{q}%')).fetchall()
    else:
        rows = db.execute('''
            SELECT v.*, c.nombre || ' ' || c.apellido AS cliente
            FROM vehiculos v LEFT JOIN clientes c ON v.cliente_id=c.id
            ORDER BY v.marca
        ''').fetchall()
    db.close()
    return render_template('vehiculos.html', vehiculos=rows, q=q)

@app.route('/vehiculos/nuevo', methods=['GET', 'POST'])
def nuevo_vehiculo():
    db = get_db()
    clientes = db.execute('SELECT id, nombre || " " || apellido AS nombre FROM clientes ORDER BY apellido').fetchall()
    if request.method == 'POST':
        marca = request.form['marca'].strip()
        modelo = request.form['modelo'].strip()
        patente = request.form['patente'].strip().upper()
        anio = request.form.get('anio', '').strip()
        motor = request.form.get('motor', '').strip()
        vin = request.form.get('vin', '').strip()
        color = request.form.get('color', '').strip()
        combustible = request.form.get('combustible', '').strip()
        cliente_id = request.form.get('cliente_id') or None
        if not marca or not modelo or not patente:
            flash('Marca, modelo y patente son obligatorios.', 'danger')
        else:
            try:
                db.execute('INSERT INTO vehiculos (marca,modelo,patente,anio,motor,vin,color,combustible,cliente_id) VALUES (?,?,?,?,?,?,?,?,?)',
                           (marca, modelo, patente, anio or None, motor or None, vin or None, color or None, combustible or None, cliente_id))
                db.commit()
                db.close()
                flash('Vehículo agregado correctamente.', 'success')
                return redirect(url_for('vehiculos'))
            except Exception as e:
                flash('La patente ya existe en el sistema.', 'danger')
    db.close()
    return render_template('vehiculo_form.html', vehiculo=None, clientes=clientes)

@app.route('/vehiculos/<int:id>/editar', methods=['GET', 'POST'])
def editar_vehiculo(id):
    db = get_db()
    vehiculo = db.execute('SELECT * FROM vehiculos WHERE id=?', (id,)).fetchone()
    clientes = db.execute('SELECT id, nombre || " " || apellido AS nombre FROM clientes ORDER BY apellido').fetchall()
    if request.method == 'POST':
        marca = request.form['marca'].strip()
        modelo = request.form['modelo'].strip()
        patente = request.form['patente'].strip().upper()
        anio = request.form.get('anio', '').strip()
        motor = request.form.get('motor', '').strip()
        vin = request.form.get('vin', '').strip()
        color = request.form.get('color', '').strip()
        combustible = request.form.get('combustible', '').strip()
        cliente_id = request.form.get('cliente_id') or None
        try:
            db.execute('UPDATE vehiculos SET marca=?,modelo=?,patente=?,anio=?,motor=?,vin=?,color=?,combustible=?,cliente_id=? WHERE id=?',
                       (marca, modelo, patente, anio or None, motor or None, vin or None, color or None, combustible or None, cliente_id, id))
            db.commit()
            db.close()
            flash('Vehículo actualizado.', 'success')
            return redirect(url_for('vehiculos'))
        except:
            flash('La patente ya existe en el sistema.', 'danger')
    db.close()
    return render_template('vehiculo_form.html', vehiculo=vehiculo, clientes=clientes)

@app.route('/vehiculos/<int:id>/eliminar', methods=['POST'])
def eliminar_vehiculo(id):
    db = get_db()
    db.execute('DELETE FROM vehiculos WHERE id=?', (id,))
    db.commit()
    db.close()
    flash('Vehículo eliminado.', 'success')
    return redirect(url_for('vehiculos'))

# ─── ÓRDENES ────────────────────────────────────────────────────────────────

@app.route('/ordenes')
def ordenes():
    q = request.args.get('q', '')
    estado = request.args.get('estado', '')
    db = get_db()
    sql = '''
        SELECT o.id, o.fecha_ingreso, o.fecha_egreso, o.estado,
               c.nombre || ' ' || c.apellido AS cliente,
               v.marca || ' ' || v.modelo AS vehiculo, v.patente
        FROM ordenes o
        JOIN clientes c ON o.cliente_id=c.id
        JOIN vehiculos v ON o.vehiculo_id=v.id
        WHERE 1=1
    '''
    params = []
    if q:
        sql += ' AND (c.nombre LIKE ? OR c.apellido LIKE ? OR v.patente LIKE ?)'
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if estado:
        sql += ' AND o.estado=?'
        params.append(estado)
    sql += ' ORDER BY o.created_at DESC'
    rows = db.execute(sql, params).fetchall()
    db.close()
    return render_template('ordenes.html', ordenes=rows, q=q, estado=estado)

@app.route('/ordenes/nueva', methods=['GET', 'POST'])
def nueva_orden():
    db = get_db()
    clientes = db.execute('SELECT id, nombre || " " || apellido AS nombre FROM clientes ORDER BY apellido').fetchall()
    vehiculos = db.execute('''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''').fetchall()
    if request.method == 'POST':
        vehiculo_id = request.form['vehiculo_id']
        cliente_id = request.form['cliente_id']
        fecha_ingreso = request.form['fecha_ingreso']
        fecha_estimada = request.form.get('fecha_estimada') or None
        fecha_egreso = request.form.get('fecha_egreso') or None
        kilometros = request.form.get('kilometros') or None
        receptor_servicio = request.form.get('receptor_servicio', '').strip()
        descripcion_trabajo = request.form.get('descripcion_trabajo', '').strip()
        diagnostico = request.form.get('diagnostico', '').strip()
        repuestos = request.form.get('repuestos', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        estado = request.form.get('estado', 'abierta')
        db.execute('''INSERT INTO ordenes
            (vehiculo_id,cliente_id,fecha_ingreso,fecha_estimada,fecha_egreso,kilometros,receptor_servicio,descripcion_trabajo,diagnostico,repuestos,observaciones,estado)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (vehiculo_id, cliente_id, fecha_ingreso, fecha_estimada, fecha_egreso, kilometros,
             receptor_servicio, descripcion_trabajo, diagnostico, repuestos, observaciones, estado))
        db.commit()
        db.close()
        flash('Orden de trabajo creada.', 'success')
        return redirect(url_for('ordenes'))
    db.close()
    return render_template('orden_form.html', orden=None, clientes=clientes, vehiculos=vehiculos)

@app.route('/ordenes/<int:id>')
def ver_orden(id):
    db = get_db()
    orden = db.execute('''
        SELECT o.*,
               c.nombre || ' ' || c.apellido AS cliente_nombre,
               c.apellido AS cliente_apellido, c.nombre AS cliente_nombre2,
               c.telefono AS cliente_tel, c.direccion AS cliente_dir,
               c.email AS cliente_email, c.dni AS cliente_dni,
               v.marca, v.modelo, v.patente, v.anio, v.motor,
               v.vin, v.color, v.combustible
        FROM ordenes o
        JOIN clientes c ON o.cliente_id=c.id
        JOIN vehiculos v ON o.vehiculo_id=v.id
        WHERE o.id=?
    ''', (id,)).fetchone()
    db.close()
    return render_template('orden_detalle.html', orden=orden)

@app.route('/ordenes/<int:id>/imprimir')
def imprimir_orden(id):
    db = get_db()
    orden = db.execute('''
        SELECT o.*,
               c.nombre || ' ' || c.apellido AS cliente_nombre,
               c.apellido AS cliente_apellido, c.nombre AS cliente_nombre2,
               c.telefono AS cliente_tel, c.direccion AS cliente_dir,
               c.email AS cliente_email, c.dni AS cliente_dni,
               v.marca, v.modelo, v.patente, v.anio, v.motor,
               v.vin, v.color, v.combustible
        FROM ordenes o
        JOIN clientes c ON o.cliente_id=c.id
        JOIN vehiculos v ON o.vehiculo_id=v.id
        WHERE o.id=?
    ''', (id,)).fetchone()
    db.close()
    return render_template('orden_print.html', orden=orden, taller=TALLER)

@app.route('/ordenes/<int:id>/editar', methods=['GET', 'POST'])
def editar_orden(id):
    db = get_db()
    orden = db.execute('SELECT * FROM ordenes WHERE id=?', (id,)).fetchone()
    clientes = db.execute('SELECT id, nombre || " " || apellido AS nombre FROM clientes ORDER BY apellido').fetchall()
    vehiculos = db.execute('''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''').fetchall()
    if request.method == 'POST':
        vehiculo_id = request.form['vehiculo_id']
        cliente_id = request.form['cliente_id']
        fecha_ingreso = request.form['fecha_ingreso']
        fecha_estimada = request.form.get('fecha_estimada') or None
        fecha_egreso = request.form.get('fecha_egreso') or None
        kilometros = request.form.get('kilometros') or None
        receptor_servicio = request.form.get('receptor_servicio', '').strip()
        descripcion_trabajo = request.form.get('descripcion_trabajo', '').strip()
        diagnostico = request.form.get('diagnostico', '').strip()
        repuestos = request.form.get('repuestos', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        estado = request.form.get('estado', 'abierta')
        db.execute('''UPDATE ordenes SET
            vehiculo_id=?,cliente_id=?,fecha_ingreso=?,fecha_estimada=?,fecha_egreso=?,kilometros=?,
            receptor_servicio=?,descripcion_trabajo=?,diagnostico=?,repuestos=?,observaciones=?,estado=?
            WHERE id=?''',
            (vehiculo_id, cliente_id, fecha_ingreso, fecha_estimada, fecha_egreso, kilometros,
             receptor_servicio, descripcion_trabajo, diagnostico, repuestos, observaciones, estado, id))
        db.commit()
        db.close()
        flash('Orden actualizada.', 'success')
        return redirect(url_for('ver_orden', id=id))
    db.close()
    return render_template('orden_form.html', orden=orden, clientes=clientes, vehiculos=vehiculos)

@app.route('/ordenes/<int:id>/cerrar', methods=['POST'])
def cerrar_orden(id):
    db = get_db()
    from datetime import date
    db.execute("UPDATE ordenes SET estado='cerrada', fecha_egreso=? WHERE id=?",
               (date.today().isoformat(), id))
    db.commit()
    db.close()
    flash('Orden cerrada.', 'success')
    return redirect(url_for('ver_orden', id=id))

@app.route('/ordenes/<int:id>/eliminar', methods=['POST'])
def eliminar_orden(id):
    db = get_db()
    db.execute('DELETE FROM ordenes WHERE id=?', (id,))
    db.commit()
    db.close()
    flash('Orden eliminada.', 'success')
    return redirect(url_for('ordenes'))

# ─── EXPORTAR EXCEL ─────────────────────────────────────────────────────────

@app.route('/exportar/ordenes')
def exportar_ordenes():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Instala openpyxl para exportar: pip install openpyxl', 'danger')
        return redirect(url_for('ordenes'))

    db = get_db()
    rows = db.execute('''
        SELECT o.id, o.fecha_ingreso, o.fecha_egreso, o.estado,
               c.nombre || ' ' || c.apellido AS cliente, c.telefono,
               v.marca, v.modelo, v.patente, v.anio,
               o.diagnostico, o.descripcion_trabajo, o.repuestos
        FROM ordenes o
        JOIN clientes c ON o.cliente_id=c.id
        JOIN vehiculos v ON o.vehiculo_id=v.id
        ORDER BY o.fecha_ingreso DESC
    ''').fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Órdenes de Trabajo'

    headers = ['N°', 'Ingreso', 'Egreso', 'Estado', 'Cliente', 'Teléfono',
               'Marca', 'Modelo', 'Patente', 'Año', 'Diagnóstico', 'Trabajo realizado', 'Repuestos']

    header_fill = PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append(list(row))

    col_widths = [6, 12, 12, 10, 25, 15, 12, 15, 12, 6, 30, 35, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='ordenes_taller.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/exportar/clientes')
def exportar_clientes():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Instala openpyxl para exportar: pip install openpyxl', 'danger')
        return redirect(url_for('clientes'))

    db = get_db()
    rows = db.execute('''
        SELECT c.id, c.nombre, c.apellido, c.direccion, c.telefono,
               COUNT(v.id) AS vehiculos
        FROM clientes c
        LEFT JOIN vehiculos v ON v.cliente_id=c.id
        GROUP BY c.id ORDER BY c.apellido
    ''').fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    headers = ['ID', 'Nombre', 'Apellido', 'Dirección', 'Teléfono', 'Vehículos']
    header_fill = PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(list(row))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='clientes_taller.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── PRESUPUESTOS ───────────────────────────────────────────────────────────

@app.route('/presupuestos')
def presupuestos():
    q = request.args.get('q', '')
    estado = request.args.get('estado', '')
    db = get_db()
    sql = '''
        SELECT p.id, p.fecha, p.estado, p.descripcion,
               c.nombre || ' ' || c.apellido AS cliente,
               v.marca || ' ' || v.modelo AS vehiculo, v.patente,
               COALESCE((SELECT SUM(cantidad * precio_unitario) FROM presupuesto_items WHERE presupuesto_id=p.id), 0) AS total
        FROM presupuestos p
        JOIN clientes c ON p.cliente_id=c.id
        LEFT JOIN vehiculos v ON p.vehiculo_id=v.id
        WHERE 1=1
    '''
    params = []
    if q:
        sql += ' AND (c.nombre LIKE ? OR c.apellido LIKE ? OR v.patente LIKE ?)'
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if estado:
        sql += ' AND p.estado=?'
        params.append(estado)
    sql += ' ORDER BY p.created_at DESC'
    rows = db.execute(sql, params).fetchall()
    db.close()
    return render_template('presupuestos.html', presupuestos=rows, q=q, estado=estado)

@app.route('/presupuestos/nuevo', methods=['GET', 'POST'])
def nuevo_presupuesto():
    db = get_db()
    clientes = db.execute('SELECT id, nombre || " " || apellido AS nombre FROM clientes ORDER BY apellido').fetchall()
    vehiculos = db.execute('''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''').fetchall()
    if request.method == 'POST':
        cliente_id = request.form['cliente_id']
        vehiculo_id = request.form.get('vehiculo_id') or None
        fecha = request.form['fecha']
        valido_hasta = request.form.get('valido_hasta') or None
        descripcion = request.form.get('descripcion', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        atendido_por = request.form.get('atendido_por', '').strip()
        condicion_venta = request.form.get('condicion_venta', '').strip()
        estado = request.form.get('estado', 'pendiente')

        cur = db.execute(
            'INSERT INTO presupuestos (cliente_id,vehiculo_id,fecha,valido_hasta,descripcion,observaciones,atendido_por,condicion_venta,estado) VALUES (?,?,?,?,?,?,?,?,?)',
            (cliente_id, vehiculo_id, fecha, valido_hasta, descripcion, observaciones, atendido_por, condicion_venta, estado)
        )
        pid = cur.lastrowid

        # Guardar ítems
        tipos = request.form.getlist('item_tipo')
        descs = request.form.getlist('item_desc')
        cantidades = request.form.getlist('item_cantidad')
        precios = request.form.getlist('item_precio')
        stocks = request.form.getlist('item_stock')
        for i, (t, d, c, p) in enumerate(zip(tipos, descs, cantidades, precios)):
            if d.strip():
                s = 1 if i < len(stocks) and stocks[i] == '1' else 0
                db.execute(
                    'INSERT INTO presupuesto_items (presupuesto_id,tipo,descripcion,cantidad,precio_unitario,en_stock) VALUES (?,?,?,?,?,?)',
                    (pid, t, d.strip(), float(c or 1), float(p or 0), s)
                )
        db.commit()
        db.close()
        flash('Presupuesto creado.', 'success')
        return redirect(url_for('ver_presupuesto', id=pid))
    db.close()
    return render_template('presupuesto_form.html', presupuesto=None, clientes=clientes, vehiculos=vehiculos, items=[])

@app.route('/presupuestos/<int:id>')
def ver_presupuesto(id):
    db = get_db()
    p = db.execute('''
        SELECT p.*,
               c.nombre || ' ' || c.apellido AS cliente_nombre,
               c.telefono AS cliente_tel, c.direccion AS cliente_dir,
               v.marca, v.modelo, v.patente, v.anio
        FROM presupuestos p
        JOIN clientes c ON p.cliente_id=c.id
        LEFT JOIN vehiculos v ON p.vehiculo_id=v.id
        WHERE p.id=?
    ''', (id,)).fetchone()
    items = db.execute('SELECT * FROM presupuesto_items WHERE presupuesto_id=? ORDER BY tipo, id', (id,)).fetchall()
    total = sum(i['cantidad'] * i['precio_unitario'] for i in items)
    db.close()
    return render_template('presupuesto_detalle.html', p=p, items=items, total=total)

@app.route('/presupuestos/<int:id>/editar', methods=['GET', 'POST'])
def editar_presupuesto(id):
    db = get_db()
    presupuesto = db.execute('SELECT * FROM presupuestos WHERE id=?', (id,)).fetchone()
    clientes = db.execute('SELECT id, nombre || " " || apellido AS nombre FROM clientes ORDER BY apellido').fetchall()
    vehiculos = db.execute('''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''').fetchall()
    items = db.execute('SELECT * FROM presupuesto_items WHERE presupuesto_id=? ORDER BY tipo, id', (id,)).fetchall()
    if request.method == 'POST':
        cliente_id = request.form['cliente_id']
        vehiculo_id = request.form.get('vehiculo_id') or None
        fecha = request.form['fecha']
        valido_hasta = request.form.get('valido_hasta') or None
        descripcion = request.form.get('descripcion', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        atendido_por = request.form.get('atendido_por', '').strip()
        condicion_venta = request.form.get('condicion_venta', '').strip()
        estado = request.form.get('estado', 'pendiente')
        db.execute('UPDATE presupuestos SET cliente_id=?,vehiculo_id=?,fecha=?,valido_hasta=?,descripcion=?,observaciones=?,atendido_por=?,condicion_venta=?,estado=? WHERE id=?',
                   (cliente_id, vehiculo_id, fecha, valido_hasta, descripcion, observaciones, atendido_por, condicion_venta, estado, id))
        db.execute('DELETE FROM presupuesto_items WHERE presupuesto_id=?', (id,))
        tipos = request.form.getlist('item_tipo')
        descs = request.form.getlist('item_desc')
        cantidades = request.form.getlist('item_cantidad')
        precios = request.form.getlist('item_precio')
        stocks = request.form.getlist('item_stock')
        for i, (t, d, c, p) in enumerate(zip(tipos, descs, cantidades, precios)):
            if d.strip():
                s = 1 if i < len(stocks) and stocks[i] == '1' else 0
                db.execute(
                    'INSERT INTO presupuesto_items (presupuesto_id,tipo,descripcion,cantidad,precio_unitario,en_stock) VALUES (?,?,?,?,?,?)',
                    (id, t, d.strip(), float(c or 1), float(p or 0), s)
                )
        db.commit()
        db.close()
        flash('Presupuesto actualizado.', 'success')
        return redirect(url_for('ver_presupuesto', id=id))
    db.close()
    return render_template('presupuesto_form.html', presupuesto=presupuesto, clientes=clientes, vehiculos=vehiculos, items=items)

@app.route('/presupuestos/<int:id>/estado/<nuevo_estado>', methods=['POST'])
def cambiar_estado_presupuesto(id, nuevo_estado):
    if nuevo_estado not in ('pendiente', 'aprobado', 'rechazado'):
        flash('Estado inválido.', 'danger')
        return redirect(url_for('ver_presupuesto', id=id))
    db = get_db()
    db.execute('UPDATE presupuestos SET estado=? WHERE id=?', (nuevo_estado, id))
    db.commit()
    db.close()
    flash(f'Presupuesto marcado como {nuevo_estado}.', 'success')
    return redirect(url_for('ver_presupuesto', id=id))

@app.route('/presupuestos/<int:id>/imprimir')
def imprimir_presupuesto(id):
    db = get_db()
    p = db.execute('''
        SELECT p.*,
               c.nombre || ' ' || c.apellido AS cliente_nombre,
               c.telefono AS cliente_tel, c.direccion AS cliente_dir,
               v.marca, v.modelo, v.patente, v.anio
        FROM presupuestos p
        JOIN clientes c ON p.cliente_id=c.id
        LEFT JOIN vehiculos v ON p.vehiculo_id=v.id
        WHERE p.id=?
    ''', (id,)).fetchone()
    items = db.execute('SELECT * FROM presupuesto_items WHERE presupuesto_id=? ORDER BY tipo, id', (id,)).fetchall()
    total = sum(i['cantidad'] * i['precio_unitario'] for i in items)
    db.close()
    return render_template('presupuesto_print.html', p=p, items=items, total=total, taller=TALLER)

@app.route('/presupuestos/<int:id>/eliminar', methods=['POST'])
def eliminar_presupuesto(id):
    db = get_db()
    db.execute('DELETE FROM presupuesto_items WHERE presupuesto_id=?', (id,))
    db.execute('DELETE FROM presupuestos WHERE id=?', (id,))
    db.commit()
    db.close()
    flash('Presupuesto eliminado.', 'success')
    return redirect(url_for('presupuestos'))

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
